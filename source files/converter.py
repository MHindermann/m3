from __future__ import annotations
from typing import List, Set, Dict, Tuple, Optional, Union
from openpyxl import load_workbook
from collections import OrderedDict

import json
from os import path

def load_codes(workbook) -> List[str]:
    """ Load all codes from workbook """

    codes = []
    wb = load_workbook(workbook)
    for ws in wb:
        for row in ws.iter_rows(min_row=7, min_col=1, max_col=1, values_only=True):
            code = row[0]
            if code is None:
                continue
            else:
                code = code.replace(" ", "")
                codes.append(code)
    return codes


DIR = path.dirname(path.abspath(__file__))
WORKBOOK = path.join(DIR, "OWC_Text.xlsx")
CODES = load_codes(WORKBOOK)
SCHEME = {"uri": "https://bartoc.org/owc/",
          "type": "skos:ConceptScheme",
          "label": "OWC Geographical Divisions"}


def main(workbook: str) -> None:
    """ Main app"""

    convert(workbook)


def load_template() -> OrderedDict:
    """ Load template """

    with open("template.json", 'r') as file:
        template = json.load(file)
    vocabulary = OrderedDict()
    vocabulary.update(template)
    return vocabulary


def convert(workbook: str) -> None:
    """ Convert workbook to JSON-LD in SKOS format """

    vocabulary = load_template()
    graph = [SCHEME]

    wb = load_workbook(workbook)
    for ws in wb:
        for row in ws.iter_rows(min_row=3497, min_col=1, max_col=3, max_row=3508, values_only=True):  # max_row for dev

            entry = OrderedDict()

            # convert code to uri and add type (fixed):
            code = row[0]
            if code is None:
                continue
            code = code.replace(" ", "")
            uri = make_uri(code)
            entry.update({"uri": uri})
            entry.update({"type": "skos:Concept"})

            # add labels:
            labels = parse(row[1])
            entry.update(labels)

            # add hierarchy:
            entry.update(make_hierarchy(code))

            graph.append(entry)

    vocabulary.update({"graph": graph})

    # print for debug
    print(json.dumps(vocabulary, indent=4, sort_keys=False))

    return vocabulary


def make_top(code: str) -> Dict:
    """ Make hierarchy for a top concept """

    codes = CODES.copy()
    broader = None
    narrower = []
    for entry in codes:
        # exclude (other) top concepts (e.g., A):
        if len(entry) < 2:
            continue
        # first symbol of entry matches:
        if code[0] is entry[0]:
            # second symbol is number (e.g, A1):
            if entry[1] in str(set(range(0, 10))):
                narrower.append({"uri": make_uri(entry)})
            # second symbol is not number and no other numbers (e.g., AA):
            elif len(entry) < 3:
                narrower.append({"uri": make_uri(entry)})
    return {"broader": broader,
            "narrower": narrower}


def make_middle(code: str, debug: int = 0) -> Dict:
    """ Make hierarchy for a middle concept """

    if debug == 1:
        print(f"middle concept with code {code}")
    codes = CODES.copy()
    broader = [{"uri": make_uri(code[0])}]
    narrower = []
    for entry in codes:
        if entry == code:
            continue
        elif code in entry and "." not in entry:
            narrower.append({"uri": make_uri(entry)})
    return {"broader": broader,
            "narrower": narrower}


def make_ojn_dot(code: str) -> Dict:
    """ Make hierarchy for OJn-dot(-dot) concepts (e.g., OJ5.11 or OJ5.11.cAbau) """

    codes = CODES.copy()
    # OJ-dot-dot concept (e.g., OJ5.11.cAbau):
    if code.count(".") == 2:
        narrower = None
        broader = [{"uri": make_uri(code.split(".")[0] + "." + code.split(".")[1])}]
    # OJ-dot concept (e.g, OJ5.11):
    else:
        broader = [{"uri": make_uri(code.split(".")[0])}]
        narrower = []
        for entry in codes:
            if entry == code:
                continue
            elif code in entry:
                narrower.append({"uri": make_uri(entry)})
        if len(narrower) == 0:
            narrower = None
    return {"broader": broader,
            "narrower": narrower}


def make_ojn(code: str) -> Dict:
    """ Make hierarchy for OJn concepts (e.g., OJ1) """

    codes = CODES.copy()
    broader = [{"uri": make_uri("OJ")}]
    narrower = []
    for entry in codes:
        if entry == code:
            continue
        elif code in entry and entry.count(".") == 1:
            narrower.append({"uri": make_uri(entry)})
    if len(narrower) == 0:
        narrower = None
    return {"broader": broader,
            "narrower": narrower}


def make_hierarchy(code: str) -> Dict:
    """ Make broader and narrower for code """

    # top concept:
    if len(code) == 1:
        return make_top(code)
    # middle concept:
    elif len(set(code).intersection(str(set(range(0, 10))))) == 0:
        return make_middle(code)
    # bottom concept:
    else:
        print(f"bottom concept with code {code}")  # debug
        # OJn special case:
        if "." in code:
            return make_ojn_dot(code)

        elif "OJ" in code:
            return make_ojn(code)
        # standard case:
        else:
            narrower = None
            broader = [{"uri": make_uri(code[:2])}]
            return {"broader": broader,
                    "narrower": narrower}


def make_uri(code: str) -> str:
    """ Convert OCW-code to URI """

    code = code.replace(" ", "")
    uri = "https://bartoc.org/ocw/" + code
    return uri


def parse(label: str) -> Dict:
    """ Parse OCW-label"""

    label_copy = label

    # prefLabel:
    label_split = label_copy.split(".", 1)
    value = label_split[0] + "."
    pref_label = {"lang": "en",
                  "value": value}

    # definition:
    if len(label_split) < 2 or label_split == "":
        value = None
    else:
        value = label_split[1]
    definition = {"lang": "en",
                  "value": value}  # text hinter dem punkt

    # altLabel:
    alt_label = None  # text in klammern oder k, evtl eher hiddenLabel

    # inScheme:
    in_scheme = None  # wenn top label, dann name des schemas

    labels = {"prefLabel": pref_label,
              "altLabel": alt_label,
              "definition": definition,
              "inScheme": in_scheme}

    return labels


main(WORKBOOK)
