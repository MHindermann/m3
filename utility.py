""" utility.py """

from typing import List
from openpyxl import load_workbook
from collections import OrderedDict
import json

def load_codes(workbook) -> List[str]:
    """ Load all codes from workbook """

    codes = []
    wb = load_workbook(workbook)
    for ws in wb:
        for row in ws.iter_rows(min_row=7, min_col=1, max_col=1, values_only=True):
            code = row[0]
            if code is None:
                continue
            else:
                code = code.replace(" ", "")
                codes.append(code)
    return codes


def load_template() -> OrderedDict:
    """ Load template """

    with open("template.json", 'r') as file:
        template = json.load(file)
    vocabulary = OrderedDict()
    vocabulary.update(template)
    return vocabulary